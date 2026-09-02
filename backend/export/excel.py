"""Exportación a Excel con FÓRMULAS VIVAS.

Un expediente técnico no admite una memoria de metrados que sea una foto de
números: el revisor abre la celda y quiere ver de dónde sale. Por eso el parcial
se escribe como `=PRODUCT(...)`, que además reproduce exactamente la regla de la
planilla — **PRODUCT ignora las celdas vacías**, igual que el motor omite las
columnas sin dato en vez de multiplicar por cero.
"""
from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

AZUL = "1D5BD8"
GRIS = "51607A"
GRIS_SUAVE = "F4F6FB"
VERDE = "0D9463"
ROJO = "D6455D"

FUENTE = "Calibri"

borde_fino = Border(*(Side(style="thin", color="DFE5EF"),) * 4)


def _titulo(hoja, texto: str, fila: int, columnas: int, tamano: int = 14) -> int:
    hoja.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=columnas)
    celda = hoja.cell(row=fila, column=1, value=texto)
    celda.font = Font(name=FUENTE, size=tamano, bold=True, color=AZUL)
    celda.alignment = Alignment(horizontal="left", vertical="center")
    hoja.row_dimensions[fila].height = 22
    return fila + 1


def _cabecera(hoja, fila: int, encabezados: list[str], anchos: list[int]) -> int:
    for i, (texto, ancho) in enumerate(zip(encabezados, anchos), start=1):
        celda = hoja.cell(row=fila, column=i, value=texto)
        celda.font = Font(name=FUENTE, size=9, bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor=GRIS)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = borde_fino
        hoja.column_dimensions[get_column_letter(i)].width = ancho
    hoja.row_dimensions[fila].height = 28
    hoja.freeze_panes = hoja.cell(row=fila + 1, column=1)
    return fila + 1


def _portada(hoja, proyecto: dict, titulo: str, filtros: str, usuario: str,
             version: str | None) -> int:
    fila = _titulo(hoja, "METRA AI", 1, 8, 16)
    fila = _titulo(hoja, titulo.upper(), fila, 8, 13)
    datos = [
        ("Proyecto", proyecto.get("nombre")),
        ("Código", proyecto.get("codigo")),
        ("Cliente", proyecto.get("cliente")),
        ("Ubicación", proyecto.get("ubicacion_texto")),
        ("Responsable", proyecto.get("responsable")),
        ("Etapa", proyecto.get("etapa")),
        ("Normativa", proyecto.get("normativa")),
        ("Moneda", proyecto.get("moneda")),
        ("Versión", version or "—"),
        ("Filtros aplicados", filtros or "Ninguno"),
        ("Generado por", usuario),
        ("Fecha de emisión", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]
    fila += 1
    for etiqueta, valor in datos:
        hoja.cell(row=fila, column=1, value=etiqueta).font = Font(name=FUENTE, size=9, bold=True)
        celda = hoja.cell(row=fila, column=2, value=valor if valor else "—")
        celda.font = Font(name=FUENTE, size=9)
        hoja.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=6)
        fila += 1
    hoja.column_dimensions["A"].width = 22
    for c in "BCDEF":
        hoja.column_dimensions[c].width = 18
    return fila + 1


def _firmas(hoja, fila: int) -> int:
    fila += 2
    hoja.cell(row=fila, column=1, value="Firmas").font = Font(name=FUENTE, size=10, bold=True)
    fila += 3
    for i, rol in enumerate(("Elaboró (metrador)", "Revisó", "Aprobó (supervisor)"), start=1):
        columna = 1 + (i - 1) * 3
        hoja.cell(row=fila, column=columna, value="_" * 28).font = Font(name=FUENTE, size=9)
        hoja.cell(row=fila + 1, column=columna, value=rol).font = Font(
            name=FUENTE, size=8, color=GRIS)
        hoja.cell(row=fila + 2, column=columna, value="Nombre / CIP / Fecha").font = Font(
            name=FUENTE, size=7, color=GRIS)
    return fila + 4


def planilla_metrados(proyecto: dict, nodos: list[dict], usuario: str,
                      version: str | None = None, filtros: str = "") -> bytes:
    """Genera el libro completo: sustento, resumen y trazabilidad."""
    libro = Workbook()

    # --- Hoja 1: planilla de sustento, con fórmulas vivas -------------------
    hoja = libro.active
    hoja.title = "Planilla de metrados"
    fila = _portada(hoja, proyecto, "Planilla de metrados", filtros, usuario, version)
    encabezados = ["ÍTEM", "DESCRIPCIÓN", "UND.", "CANT.", "N° VECES",
                   "LARGO (m)", "ANCHO (m)", "ALTO (m)", "PARCIAL", "LÁMINA",
                   "UBICACIÓN / EJE", "ORIGEN"]
    anchos = [12, 52, 8, 9, 10, 11, 11, 11, 13, 12, 24, 14]
    fila = _cabecera(hoja, fila, encabezados, anchos)

    for nodo in nodos:
        if nodo["tipo"] == "titulo":
            hoja.cell(row=fila, column=1, value=nodo["item"]).font = Font(
                name=FUENTE, size=10, bold=True)
            celda = hoja.cell(row=fila, column=2, value=nodo["descripcion"].upper())
            celda.font = Font(name=FUENTE, size=10, bold=True, color=AZUL)
            for c in range(1, len(encabezados) + 1):
                hoja.cell(row=fila, column=c).fill = PatternFill("solid", fgColor=GRIS_SUAVE)
            fila += 1
            continue

        fila_partida = fila
        hoja.cell(row=fila, column=1, value=nodo["item"]).font = Font(name=FUENTE, size=9, bold=True)
        hoja.cell(row=fila, column=2, value=nodo["descripcion"]).font = Font(
            name=FUENTE, size=9, bold=True)
        hoja.cell(row=fila, column=3, value=nodo.get("unidad")).alignment = Alignment(
            horizontal="center")
        fila += 1

        primera = fila
        for f in nodo.get("filas") or []:
            hoja.cell(row=fila, column=2, value=f.get("descripcion") or "")
            for i, clave in enumerate(("n", "veces", "largo", "ancho", "alto"), start=4):
                valor = f.get(clave)
                if valor not in (None, ""):
                    hoja.cell(row=fila, column=i, value=float(valor)).number_format = "0.00"
            signo = -1 if int(f.get("signo") or 1) < 0 else 1
            formula = f"=PRODUCT(D{fila}:H{fila})"
            if signo < 0:
                formula = f"=-PRODUCT(D{fila}:H{fila})"
            if f.get("metodo") == "formula":
                # Con fórmula libre no hay equivalente en Excel: se escribe el
                # valor y la expresión queda como comentario en la descripción.
                hoja.cell(row=fila, column=9,
                          value=float(f["parcial"]) if f.get("parcial") else None)
                hoja.cell(row=fila, column=2,
                          value=f"{f.get('descripcion') or ''}  [{f.get('sustento')}]")
            else:
                hoja.cell(row=fila, column=9, value=formula)
            hoja.cell(row=fila, column=9).number_format = "0.00"
            hoja.cell(row=fila, column=10, value=f.get("lamina") or "")
            hoja.cell(row=fila, column=11, value=f.get("eje") or "")
            hoja.cell(row=fila, column=12, value=f.get("origen") or "")
            for c in range(1, len(encabezados) + 1):
                hoja.cell(row=fila, column=c).font = Font(name=FUENTE, size=9)
                hoja.cell(row=fila, column=c).border = borde_fino
            fila += 1

        ultima = fila - 1
        celda_total = hoja.cell(row=fila, column=9)
        celda_total.value = f"=SUM(I{primera}:I{ultima})" if ultima >= primera else 0
        celda_total.font = Font(name=FUENTE, size=9, bold=True)
        celda_total.number_format = "0.00"
        celda_total.fill = PatternFill("solid", fgColor="EAF0FF")
        hoja.cell(row=fila, column=2, value="TOTAL").font = Font(
            name=FUENTE, size=9, bold=True, color=GRIS)
        hoja.cell(row=fila, column=3, value=nodo.get("unidad")).font = Font(
            name=FUENTE, size=9, bold=True)
        hoja.cell(row=fila_partida, column=9, value=f"=I{fila}").number_format = "0.00"
        fila += 2

    _firmas(hoja, fila)

    # --- Hoja 2: resumen de metrados ----------------------------------------
    resumen = libro.create_sheet("Resumen de metrados")
    fila = _portada(resumen, proyecto, "Resumen de metrados", filtros, usuario, version)
    fila = _cabecera(resumen, fila, ["ÍTEM", "CÓDIGO", "DESCRIPCIÓN", "UND.", "METRADO",
                                     "ESPECIALIDAD", "ESTADO"],
                     [12, 14, 60, 8, 14, 20, 12])
    for nodo in nodos:
        if nodo["tipo"] != "partida":
            continue
        resumen.cell(row=fila, column=1, value=nodo["item"])
        resumen.cell(row=fila, column=2, value=nodo.get("codigo") or "")
        resumen.cell(row=fila, column=3, value=nodo["descripcion"])
        resumen.cell(row=fila, column=4, value=nodo.get("unidad"))
        resumen.cell(row=fila, column=5,
                     value=float(nodo.get("metrado") or 0)).number_format = "0.00"
        resumen.cell(row=fila, column=6, value=nodo.get("especialidad"))
        resumen.cell(row=fila, column=7, value=nodo.get("estado"))
        for c in range(1, 8):
            resumen.cell(row=fila, column=c).font = Font(name=FUENTE, size=9)
            resumen.cell(row=fila, column=c).border = borde_fino
        fila += 1
    _firmas(resumen, fila)

    # --- Hoja 3: trazabilidad ------------------------------------------------
    trazas = libro.create_sheet("Sustento y trazabilidad")
    fila = _cabecera(trazas, 1, ["ÍTEM", "PARTIDA", "FILA", "CÁLCULO", "PARCIAL",
                                 "ORIGEN", "LÁMINA", "RESPONSABLE", "FECHA", "OBSERVACIÓN"],
                     [12, 40, 30, 34, 12, 14, 12, 18, 12, 34])
    for nodo in nodos:
        if nodo["tipo"] != "partida":
            continue
        for f in nodo.get("filas") or []:
            valores = [nodo["item"], nodo["descripcion"], f.get("descripcion") or "",
                       f.get("sustento") or "", f.get("parcial") or "",
                       f.get("origen") or "", f.get("lamina") or "",
                       f.get("responsable") or "", f.get("fecha") or "",
                       f.get("observacion") or f.get("supuesto") or f.get("error") or ""]
            for c, v in enumerate(valores, start=1):
                celda = trazas.cell(row=fila, column=c, value=v)
                celda.font = Font(name=FUENTE, size=8)
                celda.border = borde_fino
                celda.alignment = Alignment(wrap_text=c in (2, 3, 4, 10), vertical="top")
            fila += 1

    salida = io.BytesIO()
    libro.save(salida)
    return salida.getvalue()


def presupuesto(proyecto: dict, nodos: list[dict], resumen_costos: dict,
                usuario: str, version: str | None = None) -> bytes:
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Presupuesto"
    moneda = proyecto.get("moneda", "PEN")
    fila = _portada(hoja, proyecto, "Presupuesto de obra", "", usuario, version)
    fila = _cabecera(hoja, fila, ["ÍTEM", "CÓDIGO", "DESCRIPCIÓN", "UND.", "METRADO",
                                  f"P.U. ({moneda})", f"PARCIAL ({moneda})"],
                     [12, 14, 56, 8, 13, 14, 16])

    primera = fila
    for nodo in nodos:
        es_titulo = nodo["tipo"] == "titulo"
        hoja.cell(row=fila, column=1, value=nodo["item"])
        hoja.cell(row=fila, column=2, value=nodo.get("codigo") or "")
        hoja.cell(row=fila, column=3,
                  value=nodo["descripcion"].upper() if es_titulo else nodo["descripcion"])
        if not es_titulo:
            hoja.cell(row=fila, column=4, value=nodo.get("unidad"))
            hoja.cell(row=fila, column=5,
                      value=float(nodo.get("metrado") or 0)).number_format = "0.00"
            hoja.cell(row=fila, column=6,
                      value=float(nodo.get("precio_unitario") or 0)).number_format = "0.00"
            hoja.cell(row=fila, column=7, value=f"=E{fila}*F{fila}").number_format = "#,##0.00"
        else:
            hoja.cell(row=fila, column=7,
                      value=float(nodo.get("parcial") or 0)).number_format = "#,##0.00"
        for c in range(1, 8):
            celda = hoja.cell(row=fila, column=c)
            celda.font = Font(name=FUENTE, size=9, bold=es_titulo,
                              color=AZUL if es_titulo else "000000")
            celda.border = borde_fino
            if es_titulo:
                celda.fill = PatternFill("solid", fgColor=GRIS_SUAVE)
        fila += 1

    fila += 1
    pie = [
        ("COSTO DIRECTO", resumen_costos["costo_directo"], True),
        (f"Gastos generales ({resumen_costos['gastos_generales_pct']}%)",
         resumen_costos["gastos_generales"], False),
        (f"Utilidad ({resumen_costos['utilidad_pct']}%)", resumen_costos["utilidad"], False),
        ("SUBTOTAL", resumen_costos["subtotal"], True),
        (f"{resumen_costos['nombre_impuesto']} ({resumen_costos['impuesto_pct']}%)",
         resumen_costos["impuesto"], False),
        ("TOTAL DE PRESUPUESTO", resumen_costos["total"], True),
    ]
    for etiqueta, valor, fuerte in pie:
        hoja.cell(row=fila, column=3, value=etiqueta).font = Font(
            name=FUENTE, size=10 if fuerte else 9, bold=fuerte)
        celda = hoja.cell(row=fila, column=7, value=float(valor))
        celda.number_format = "#,##0.00"
        celda.font = Font(name=FUENTE, size=10 if fuerte else 9, bold=fuerte,
                          color=VERDE if etiqueta.startswith("TOTAL DE") else "000000")
        fila += 1

    _firmas(hoja, fila)
    salida = io.BytesIO()
    libro.save(salida)
    return salida.getvalue()


def generico(titulo: str, proyecto: dict, encabezados: list[str], filas: list[list],
             usuario: str, anchos: list[int] | None = None) -> bytes:
    """Reporte tabular simple: insumos, acero, observaciones, comparación."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = titulo[:31]
    fila = _portada(hoja, proyecto, titulo, "", usuario, None)
    fila = _cabecera(hoja, fila, encabezados,
                     anchos or [max(12, min(50, len(h) + 8)) for h in encabezados])
    for datos in filas:
        for c, v in enumerate(datos, start=1):
            celda = hoja.cell(row=fila, column=c,
                              value=float(v) if isinstance(v, (int, float)) else v)
            celda.font = Font(name=FUENTE, size=9)
            celda.border = borde_fino
            if isinstance(v, (int, float)):
                celda.number_format = "#,##0.00"
        fila += 1
    _firmas(hoja, fila)
    salida = io.BytesIO()
    libro.save(salida)
    return salida.getvalue()
