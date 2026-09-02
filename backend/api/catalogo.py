"""Catálogo de partidas e insumos: biblioteca normativa + biblioteca propia."""
from __future__ import annotations

import csv
import io

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from pydantic import BaseModel, Field
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from .. import audit, security
from ..db import obtener_sesion
from ..models import Insumo, PartidaCatalogo, Precio, Usuario
from ..motor import especialidades, normas
from ..motor.unidades import existe as unidad_existe

router = APIRouter(tags=["catálogo"])


def _publico(p: PartidaCatalogo) -> dict:
    return {
        "id": p.id, "codigo": p.codigo, "descripcion": p.descripcion,
        "unidad": p.unidad, "especialidad": p.especialidad,
        "color": especialidades.color(p.especialidad),
        "capitulo": p.capitulo, "formula": p.formula,
        "plantilla_formula": p.plantilla_formula,
        "regla_medicion": p.regla_medicion, "norma": p.norma, "pais": p.pais,
        "desperdicio_pct": p.desperdicio_pct, "rendimiento": p.rendimiento,
        "cuadrilla": p.cuadrilla, "favorita": p.favorita,
        "verificado": p.verificado, "fuente": p.fuente,
        "etiquetas": p.etiquetas or [],
        "propia": p.empresa_id is not None,
        "familia_descuento": normas.familia_por_defecto(p.especialidad, p.descripcion),
    }


@router.get("/catalogo/partidas")
def buscar(q: str | None = None, especialidad: str | None = None,
           solo_verificadas: bool = False, favoritas: bool = False,
           pagina: int = Query(default=1, ge=1), por_pagina: int = Query(default=60, ge=1, le=300),
           db: Session = Depends(obtener_sesion),
           usuario: Usuario = Depends(security.usuario_actual)):
    consulta = select(PartidaCatalogo).where(
        or_(PartidaCatalogo.empresa_id.is_(None),
            PartidaCatalogo.empresa_id == usuario.empresa_id))
    if q:
        patron = f"%{q.strip().lower()}%"
        consulta = consulta.where(or_(
            func.lower(PartidaCatalogo.descripcion).like(patron),
            func.lower(PartidaCatalogo.codigo).like(patron)))
    if especialidad:
        consulta = consulta.where(PartidaCatalogo.especialidad == especialidad)
    if solo_verificadas:
        consulta = consulta.where(PartidaCatalogo.verificado.is_(True))
    if favoritas:
        consulta = consulta.where(PartidaCatalogo.favorita.is_(True))

    total = db.scalar(select(func.count()).select_from(consulta.subquery())) or 0
    filas = list(db.scalars(consulta.order_by(PartidaCatalogo.codigo)
                            .offset((pagina - 1) * por_pagina).limit(por_pagina)))
    return {"partidas": [_publico(p) for p in filas], "total": total,
            "pagina": pagina, "por_pagina": por_pagina}


@router.get("/catalogo/resumen")
def resumen_catalogo(db: Session = Depends(obtener_sesion),
                     usuario: Usuario = Depends(security.usuario_actual)):
    totales = dict(db.execute(
        select(PartidaCatalogo.especialidad, func.count(PartidaCatalogo.id))
        .group_by(PartidaCatalogo.especialidad)).all())
    verificadas_por = dict(db.execute(
        select(PartidaCatalogo.especialidad, func.count(PartidaCatalogo.id))
        .where(PartidaCatalogo.verificado.is_(True))
        .group_by(PartidaCatalogo.especialidad)).all())
    por_especialidad = []
    for clave, total in totales.items():
        e = especialidades.POR_CLAVE.get(clave, {})
        por_especialidad.append({
            "clave": clave, "nombre": e.get("nombre", clave), "color": e.get("color", "#8593ab"),
            "total": total, "verificadas": verificadas_por.get(clave, 0),
        })
    por_especialidad.sort(key=lambda x: especialidades.POR_CLAVE.get(x["clave"], {}).get("orden", 99))
    return {
        "por_especialidad": por_especialidad,
        "total": db.scalar(select(func.count(PartidaCatalogo.id))) or 0,
        "verificadas": db.scalar(select(func.count(PartidaCatalogo.id))
                                 .where(PartidaCatalogo.verificado.is_(True))) or 0,
    }


class DatosPartida(BaseModel):
    codigo: str = Field(min_length=1, max_length=40)
    descripcion: str = Field(min_length=3, max_length=400)
    unidad: str
    especialidad: str = "arquitectura"
    formula: str | None = None
    plantilla_formula: str | None = None
    regla_medicion: str | None = None
    desperdicio_pct: str | None = None
    rendimiento: str | None = None
    cuadrilla: str | None = None
    fuente: str | None = None


@router.post("/catalogo/partidas", status_code=201)
def crear_partida(datos: DatosPartida, db: Session = Depends(obtener_sesion),
                  usuario: Usuario = Depends(security.usuario_actual)):
    security.exigir(db, usuario, None, "crear")
    if not unidad_existe(datos.unidad):
        raise HTTPException(400, f"Unidad desconocida: {datos.unidad}")
    p = PartidaCatalogo(empresa_id=usuario.empresa_id, verificado=False,
                        etiquetas=["propia"], **datos.model_dump())
    db.add(p)
    audit.registrar(db, accion="crear", entidad="partida_catalogo", entidad_id=p.id,
                    resumen=f"Partida propia: {p.codigo} {p.descripcion[:80]}", usuario=usuario)
    db.commit()
    return {"partida": _publico(p)}


@router.put("/catalogo/partidas/{partida_id}")
def actualizar_partida(partida_id: str, datos: dict, db: Session = Depends(obtener_sesion),
                       usuario: Usuario = Depends(security.usuario_actual)):
    p = db.get(PartidaCatalogo, partida_id)
    if not p:
        raise HTTPException(404, "Partida no encontrada.")
    if p.empresa_id is None and "favorita" not in datos:
        raise HTTPException(
            403,
            "Las partidas del catálogo normativo no se editan: su texto es el de la norma. "
            "Duplíquela para crear una versión propia.")
    permitidos = {"favorita"} if p.empresa_id is None else {
        "codigo", "descripcion", "unidad", "especialidad", "formula", "plantilla_formula",
        "regla_medicion", "desperdicio_pct", "rendimiento", "cuadrilla", "fuente", "favorita"}
    for campo, valor in datos.items():
        if campo in permitidos:
            setattr(p, campo, valor)
    db.commit()
    return {"partida": _publico(p)}


@router.post("/catalogo/partidas/{partida_id}/duplicar", status_code=201)
def duplicar_partida(partida_id: str, db: Session = Depends(obtener_sesion),
                     usuario: Usuario = Depends(security.usuario_actual)):
    original = db.get(PartidaCatalogo, partida_id)
    if not original:
        raise HTTPException(404, "Partida no encontrada.")
    copia = PartidaCatalogo(
        empresa_id=usuario.empresa_id, codigo=f"{original.codigo}-P",
        descripcion=original.descripcion, unidad=original.unidad,
        especialidad=original.especialidad, capitulo=original.capitulo,
        formula=original.formula, plantilla_formula=original.plantilla_formula,
        regla_medicion=original.regla_medicion, norma=original.norma, pais=original.pais,
        verificado=False, fuente=f"Copia de {original.codigo}",
        etiquetas=["copia editable"],
    )
    db.add(copia)
    db.commit()
    return {"partida": _publico(copia)}


@router.delete("/catalogo/partidas/{partida_id}")
def eliminar_partida(partida_id: str, db: Session = Depends(obtener_sesion),
                     usuario: Usuario = Depends(security.usuario_actual)):
    p = db.get(PartidaCatalogo, partida_id)
    if not p:
        raise HTTPException(404, "Partida no encontrada.")
    if p.empresa_id is None:
        raise HTTPException(403, "No se elimina una partida del catálogo normativo.")
    db.delete(p)
    db.commit()
    return {"ok": True}


@router.post("/catalogo/importar")
async def importar_catalogo(archivo: UploadFile = File(...),
                            db: Session = Depends(obtener_sesion),
                            usuario: Usuario = Depends(security.usuario_actual)):
    """Importa un catálogo propio desde CSV o Excel (columnas: codigo, descripcion, unidad...)."""
    security.exigir(db, usuario, None, "crear")
    contenido = await archivo.read()
    nombre = (archivo.filename or "").lower()

    if nombre.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        libro = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        hoja = libro.active
        filas = [[("" if c is None else str(c)) for c in fila]
                 for fila in hoja.iter_rows(values_only=True)]
    else:
        texto = contenido.decode("utf-8-sig", errors="replace")
        muestra = texto[:2000]
        try:
            dialecto = csv.Sniffer().sniff(muestra, delimiters=",;\t|")
        except csv.Error:
            dialecto = csv.excel
        filas = [f for f in csv.reader(io.StringIO(texto), dialecto)]

    if not filas:
        raise HTTPException(400, "El archivo está vacío.")

    cabecera = [str(c).strip().lower() for c in filas[0]]
    requeridas = {"codigo", "descripcion", "unidad"}
    if not requeridas.issubset(set(cabecera)):
        raise HTTPException(
            400,
            "El archivo debe tener al menos las columnas: codigo, descripcion, unidad. "
            f"Se encontraron: {', '.join(cabecera[:12])}")

    indice = {c: i for i, c in enumerate(cabecera)}
    creadas, errores = 0, []
    for n, fila in enumerate(filas[1:], start=2):
        def celda(nombre_col):
            i = indice.get(nombre_col)
            return (str(fila[i]).strip() if i is not None and i < len(fila) else "") or None

        codigo, descripcion, unidad = celda("codigo"), celda("descripcion"), celda("unidad")
        if not (codigo and descripcion and unidad):
            errores.append({"fila": n, "motivo": "Faltan código, descripción o unidad."})
            continue
        if not unidad_existe(unidad):
            errores.append({"fila": n, "motivo": f"Unidad desconocida: {unidad}"})
            continue
        db.add(PartidaCatalogo(
            empresa_id=usuario.empresa_id, codigo=codigo, descripcion=descripcion,
            unidad=unidad,
            especialidad=especialidades.normalizar(celda("especialidad")),
            formula=celda("formula"), regla_medicion=celda("regla_medicion"),
            desperdicio_pct=celda("desperdicio"), rendimiento=celda("rendimiento"),
            cuadrilla=celda("cuadrilla"), verificado=False,
            fuente=f"Importado de {archivo.filename}",
            etiquetas=["importado"],
        ))
        creadas += 1

    audit.registrar(db, accion="importar", entidad="partida_catalogo",
                    resumen=f"{creadas} partidas importadas de {archivo.filename}",
                    usuario=usuario)
    db.commit()
    return {"creadas": creadas, "errores": errores[:50], "total_errores": len(errores)}


# --------------------------------------------------------------------------- #
# Insumos y precios
# --------------------------------------------------------------------------- #

@router.get("/catalogo/insumos")
def listar_insumos(q: str | None = None, tipo: str | None = None,
                   db: Session = Depends(obtener_sesion),
                   usuario: Usuario = Depends(security.usuario_actual)):
    consulta = select(Insumo).where(or_(Insumo.empresa_id.is_(None),
                                        Insumo.empresa_id == usuario.empresa_id))
    if q:
        consulta = consulta.where(func.lower(Insumo.descripcion).like(f"%{q.lower()}%"))
    if tipo:
        consulta = consulta.where(Insumo.tipo == tipo)
    filas = list(db.scalars(consulta.order_by(Insumo.tipo, Insumo.descripcion)))
    return {"insumos": [{
        "id": i.id, "codigo": i.codigo, "descripcion": i.descripcion,
        "unidad": i.unidad, "tipo": i.tipo, "grupo": i.grupo,
        "precio_referencial": i.precio_referencial, "moneda": i.moneda,
    } for i in filas],
        "tipos": {"MO": "Mano de obra", "MAT": "Materiales",
                  "EQ": "Equipos", "SC": "Subcontratos"}}


class DatosPrecio(BaseModel):
    insumo_id: str
    valor: str
    fecha: str
    moneda: str = "PEN"
    lugar: str | None = None
    fuente: str | None = None
    proyecto_id: str | None = None


@router.post("/catalogo/precios", status_code=201)
def registrar_precio(datos: DatosPrecio, db: Session = Depends(obtener_sesion),
                     usuario: Usuario = Depends(security.usuario_actual)):
    if not db.get(Insumo, datos.insumo_id):
        raise HTTPException(404, "Insumo no encontrado.")
    p = Precio(**datos.model_dump())
    db.add(p)
    db.commit()
    return {"precio_id": p.id}


@router.get("/catalogo/insumos/{insumo_id}/precios")
def historico_precios(insumo_id: str, db: Session = Depends(obtener_sesion),
                      usuario: Usuario = Depends(security.usuario_actual)):
    filas = list(db.scalars(select(Precio).where(Precio.insumo_id == insumo_id)
                            .order_by(Precio.fecha.desc())))
    return {"precios": [{"id": p.id, "valor": p.valor, "moneda": p.moneda,
                         "fecha": p.fecha, "lugar": p.lugar, "fuente": p.fuente}
                        for p in filas]}
