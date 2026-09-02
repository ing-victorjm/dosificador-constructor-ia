"""
Metrados - Web App (Flask localhost)
Sirve en http://localhost:5050
"""
import sys
import io
from pathlib import Path
from unittest.mock import MagicMock

# Mock PySide6 para que geo3d (que importa vista3d.py) pueda cargarse sin Qt
for _mod in ["PySide6", "PySide6.QtCore", "PySide6.QtGui", "PySide6.QtWidgets"]:
    if _mod not in sys.modules:
        sys.modules[_mod] = MagicMock()

# Agrega el modulo de dosificacion al path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from flask import Flask, render_template, request, jsonify, send_file
from dosificacion_concreto import modelo, baldes, mortero, estribos, geo3d, elementos
from dosificacion_concreto import exportar_excel

app = Flask(__name__, template_folder="templates", static_folder="static")
app.config["TEMPLATES_AUTO_RELOAD"] = True
try:
    app.json.ensure_ascii = False          # Flask >= 2.3
except AttributeError:
    app.config["JSON_AS_ASCII"] = False    # Flask < 2.3


# ── Pagina principal ────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")


# ── API: Dosificacion ───────────────────────────────────────────────────────
@app.route("/api/dosificar", methods=["POST"])
def api_dosificar():
    d = request.json
    try:
        req = modelo.calcular(
            fc=float(d["fc"]),
            volumen_m3=float(d["volumen_m3"]),
            desperdicio_pct=float(d.get("desperdicio_pct", 5.0)),
            peso_bolsa=float(d.get("peso_bolsa", 42.5)),
        )
        dos = req.dosificacion
        return jsonify({
            "ok": True,
            "cemento_bolsas": round(req.cemento_bolsas, 2),
            "cemento_kg": round(req.cemento_kg, 1),
            "arena_m3": round(req.arena_m3, 3),
            "piedra_m3": round(req.piedra_m3, 3),
            "agua_m3": round(req.agua_m3, 3),
            "agua_litros": round(req.agua_litros, 1),
            "dosificacion": {
                "ac": dos.a_c,
                "slump": dos.slump_pulg,
                "tmax": dos.tmax_pulg,
                "proporcion": dos.dosificacion_volumen,
                "interpolado": dos.interpolado,
            },
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


# ── API: Baldes ─────────────────────────────────────────────────────────────
@app.route("/api/baldes", methods=["POST"])
def api_baldes():
    d = request.json
    try:
        req = modelo.calcular(
            fc=float(d["fc"]),
            volumen_m3=float(d["volumen_m3"]),
            desperdicio_pct=float(d.get("desperdicio_pct", 5.0)),
            peso_bolsa=float(d.get("peso_bolsa", 42.5)),
        )
        vol_balde = float(d.get("vol_balde", 0.018))
        peso_bolsa = float(d.get("peso_bolsa", 42.5))
        bal = baldes.dosificar_por_baldes(req.dosificacion, req, vol_balde, peso_bolsa)
        pedir = baldes.material_a_pedir(req)
        recom = baldes.recomendaciones_obra(req)
        palabras = baldes.palabras_maestro(req, bal, peso_bolsa)
        return jsonify({
            "ok": True,
            "baldes_arena": bal.baldes_arena,
            "baldes_piedra": bal.baldes_piedra,
            "agua_litros_bolsa": bal.agua_litros_bolsa,
            "pedir": pedir,
            "recomendaciones": recom,
            "palabras": palabras,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


# ── API: Mortero ────────────────────────────────────────────────────────────
@app.route("/api/mortero/asentado", methods=["POST"])
def api_mortero_asentado():
    d = request.json
    try:
        lad = mortero.LADRILLOS[int(d.get("ladrillo_idx", 0))]
        r = mortero.asentado(
            float(d["area"]),
            lad,
            d.get("aparejo", "Soga"),
            float(d.get("junta_cm", 1.5)),
            d.get("proporcion", "1:4"),
        )
        return jsonify({
            "ok": True,
            "ladrillos": r.ladrillos,
            "ladrillos_por_m2": r.ladrillos_por_m2,
            "cemento_bolsas": round(r.cemento_bolsas, 2),
            "arena_m3": round(r.arena_m3, 3),
            "volumen_m3": round(r.volumen_mortero_m3, 4),
            "detalle": r.detalle,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


@app.route("/api/mortero/tarrajeo", methods=["POST"])
def api_mortero_tarrajeo():
    d = request.json
    try:
        area = float(d["area"])
        t = mortero.tarrajeo(
            area,
            float(d.get("espesor_cm", 1.5)),
            d.get("proporcion", "1:5"),
            caras=int(d.get("caras", 1)),
        )
        m2_por_bolsa = round(area / t.cemento_bolsas, 1) if t.cemento_bolsas > 0 else 0
        return jsonify({
            "ok": True,
            "cemento_bolsas": round(t.cemento_bolsas, 2),
            "arena_m3": round(t.arena_m3, 3),
            "volumen_m3": round(t.volumen_mortero_m3, 4),
            "m2_por_bolsa": m2_por_bolsa,
            "detalle": t.detalle,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


# ── API: Estribos ───────────────────────────────────────────────────────────
@app.route("/api/estribos", methods=["POST"])
def api_estribos():
    d = request.json
    try:
        r = estribos.calcular(
            h_cm=float(d["h_cm"]),
            a_cm=float(d["a_cm"]),
            b_cm=float(d["b_cm"]),
            de_mm=float(d["de_mm"]),
            db_mm=float(d["db_mm"]),
            rec_cm=float(d.get("rec_cm", 4.0)),
            s1_manual=float(d["s1"]) if d.get("s1") else None,
            s2_manual=float(d["s2"]) if d.get("s2") else None,
            lo_manual=float(d["lo"]) if d.get("lo") else None,
            # La E.060 da separaciones distintas segun el sistema estructural:
            # 21.6.4 para porticos y dual II, 21.4.5 para muros y dual I.
            sistema=(d.get("sistema") or estribos.PORTICOS),
        )
        return jsonify({
            "ok": True,
            "lo_cm": r.lo_cm,
            "s1_cm": r.s1_cm,
            "s2_cm": r.s2_cm,
            "n_conf_inf": r.n_conf_inf,
            "n_central": r.n_central,
            "n_conf_sup": r.n_conf_sup,
            "n_total": r.n_total,
            "long_estribo_cm": r.long_estribo_cm,
            "peso_total_kg": r.peso_total_kg,
            "kg_por_ml": r.kg_por_ml,
            "detalle": r.detalle,
            "posiciones_cm": r.posiciones_cm,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


# ── API: Geo3D (malla Three.js) ─────────────────────────────────────────────
@app.route("/api/geo3d", methods=["POST"])
def api_geo3d():
    d = request.json
    try:
        malla = geo3d.construir_malla(d["tipo"], d["valores"])
        return jsonify({"ok": True, **malla})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


# ── API: Metrado ─────────────────────────────────────────────────────────────
@app.route("/api/metrado/calcular", methods=["POST"])
def api_metrado_calcular():
    d = request.json
    try:
        tipo_clave = d["tipo"]
        valores = {k: float(v) for k, v in d["valores"].items()}
        cantidad = int(d.get("cantidad", 1))
        tipo = next((t for t in elementos.TIPOS if t.clave == tipo_clave), None)
        if not tipo:
            return jsonify({"ok": False, "error": f"Tipo desconocido: {tipo_clave}"}), 400
        vol_unit = tipo.formula(valores)
        vol_total = vol_unit * cantidad
        # El ACERO no se calcula aqui. La Norma Tecnica de Metrados para Obras de
        # Edificacion (R.D. 073-2010-VIVIENDA/VMCS-DNC), seccion OE.2.3, exige
        # computar "el peso total del fierro indicado en los planos", sumando la
        # longitud de cada barra con sus ganchos, dobleces y traslapes, agrupada
        # por diametros iguales y multiplicada por su peso en kg/m.
        # La expresion "kg/m3" no aparece en la norma: multiplicar el volumen de
        # concreto por un ratio es predimensionamiento, no metrado, y ademas no
        # puede respetar la regla de arranques (la zapata los excluye, la columna
        # los incluye). El acero se metra en la seccion "Acero - Despiece".
        enc_unit = elementos.encofrado_m2(tipo_clave, valores)
        exc_unit = elementos.excavacion_m3(tipo_clave, valores)
        return jsonify({
            "ok": True,
            "vol_unit": round(vol_unit, 4),
            "vol_total": round(vol_total, 4),
            "ayuda": tipo.ayuda,
            "encofrado_m2": round(enc_unit * cantidad, 2),
            "excavacion_m3": round(exc_unit * cantidad, 3),
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


# ── API: Exportar Excel Dosificacion ─────────────────────────────────────────
@app.route("/api/exportar_dosificacion", methods=["POST"])
def api_exportar_dosificacion():
    d = request.json
    try:
        req = modelo.calcular(
            fc=float(d["fc"]),
            volumen_m3=float(d["volumen_m3"]),
            desperdicio_pct=float(d.get("desperdicio_pct", 5.0)),
            peso_bolsa=float(d.get("peso_bolsa", 42.5)),
        )
        buf = io.BytesIO()
        exportar_excel.exportar(buf, req, responsable=d.get("responsable", ""), presupuesto=None)
        buf.seek(0)
        return send_file(buf, download_name="dosificacion.xlsx", as_attachment=True,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


# ── API: Exportar Excel Metrado ─────────────────────────────────────────────
@app.route("/api/exportar_metrado", methods=["POST"])
def api_exportar_metrado():
    d = request.json or {}
    items = d.get("items", [])
    proyecto = (d.get("proyecto") or "").strip() or "Proyecto sin nombre"
    # El acero llega del DESPIECE (planilla real por diametro), no del volumen.
    # despiece: [{"nombre": "1/2\"", "kg_m": 0.994, "kg_varilla": 8.95, "kg": 123.4}, ...]
    despiece = d.get("despiece") or []
    desp_pct = float(d.get("desperdicio_acero") or 0)
    # kg NETO metrado (sin desperdicio: la norma lo manda al analisis de precios).
    acero_kg = sum(float(x.get("kg") or 0) for x in despiece) if despiece else None
    try:
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Metrado"
        CYAN, DARK, GRIS = "0E7490", "0F172A", "64748B"
        thin = Side(style="thin", color="CBD5E1")
        borde = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws["A1"] = "CONSTRUCTOR IA — METRADO DE CONCRETO"
        ws["A1"].font = Font(bold=True, size=14, color=DARK)
        ws["A2"] = f"Proyecto: {proyecto}"
        ws["A2"].font = Font(size=11, color=GRIS)
        ws["A3"] = "Generado: " + datetime.now().strftime("%d/%m/%Y %H:%M") + " · NTE E.060 / ACI 318"
        ws["A3"].font = Font(size=9, color=GRIS, italic=True)

        # Sin columna de acero: el acero se metra por despiece (OE.2.3), no se
        # deriva del volumen. Exportarlo aqui daba un kilaje sin sustento.
        headers = ["#", "Elemento", "Dimensiones", "Cant.", "Vol. unit (m³)",
                   "Vol. total (m³)", "Encofrado (m²)", "Excavación (m³)"]
        HR = 5
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=HR, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor=CYAN)
            cell.border = borde
            cell.alignment = Alignment(horizontal="center", vertical="center")

        tv = te = tx = 0.0
        for i, it in enumerate(items, 1):
            r = HR + i
            vals = [i, it.get("tipo", ""), it.get("dims", ""), it.get("cantidad", 1),
                    round(float(it.get("vu", 0)), 4), round(float(it.get("vt", 0)), 4),
                    round(float(it.get("encof", 0)), 2), round(float(it.get("exc", 0)), 3)]
            tv += vals[5]; te += vals[6]; tx += vals[7]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = borde
                cell.font = Font(size=10)
                if c >= 4:
                    cell.alignment = Alignment(horizontal="right")
                    if c >= 5:
                        cell.number_format = "0.000" if c in (5, 6, 8) else "0.00"
        rT = HR + len(items) + 1
        tot_vals = ["", "TOTAL", "", "", "", round(tv, 3), round(te, 2), round(tx, 3)]
        for c, v in enumerate(tot_vals, 1):
            cell = ws.cell(row=rT, column=c, value=v)
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill("solid", fgColor="E8F4F8")
            cell.border = borde
            if c >= 5:
                cell.alignment = Alignment(horizontal="right")

        # Hoja de acero SOLO si hay despiece. Antes esta hoja repartia el kilaje
        # derivado del volumen entre todos los diametros "como si usaras uno solo":
        # eso no es un despiece y no servia para comprar.
        if despiece:
            ws2 = wb.create_sheet("Acero")
            ws2["A1"] = "ACERO — planilla por diámetro"
            ws2["A1"].font = Font(bold=True, size=12, color=DARK)
            ws2["A2"] = ("Metrado NETO segun OE.2.3 (longitud con ganchos, dobleces y traslapes "
                         "× kg/m, agrupado por diámetro). El desperdicio NO va en el metrado: "
                         "la norma lo manda al análisis de precios unitarios.")
            ws2["A2"].font = Font(size=9, color=GRIS, italic=True)
            cab = ["Diámetro", "kg/m", "Peso metrado (kg)",
                   f"Para compra +{desp_pct:g}% (kg)", "Varillas 9 m"]
            for c, h in enumerate(cab, 1):
                cell = ws2.cell(row=4, column=c, value=h)
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.fill = PatternFill("solid", fgColor=CYAN)
                cell.border = borde
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            tot_neto = tot_compra = 0.0
            for j, x in enumerate(despiece):
                r = 5 + j
                kg = float(x.get("kg") or 0)
                kgv = float(x.get("kg_varilla") or 0)
                kg_c = kg * (1 + desp_pct / 100.0)
                tot_neto += kg; tot_compra += kg_c
                n = int(-(-kg_c // kgv)) if kgv > 0 else 0
                for c, v in enumerate([x.get("nombre", ""), x.get("kg_m", 0),
                                       round(kg, 1), round(kg_c, 1), n], 1):
                    cell = ws2.cell(row=r, column=c, value=v)
                    cell.border = borde
                    if c > 1:
                        cell.alignment = Alignment(horizontal="right")
            rt = 5 + len(despiece)
            for c, v in enumerate(["TOTAL", "", round(tot_neto, 1), round(tot_compra, 1), ""], 1):
                cell = ws2.cell(row=rt, column=c, value=v)
                cell.font = Font(bold=True, size=10)
                cell.fill = PatternFill("solid", fgColor="E8F4F8")
                cell.border = borde
                if c > 1:
                    cell.alignment = Alignment(horizontal="right")
            for c, w in enumerate([16, 10, 18, 18, 13], 1):
                ws2.column_dimensions[chr(64 + c)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, download_name="metrado_constructor_ia.xlsx", as_attachment=True,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


# ── API: info estatica ──────────────────────────────────────────────────────
@app.route("/api/opciones")
def api_opciones():
    return jsonify({
        "ladrillos": [{"idx": i, "nombre": l[0], "largo": l[1], "ancho": l[2], "alto": l[3]}
                      for i, l in enumerate(mortero.LADRILLOS)],
        "aparejos": list(mortero.APAREJOS.keys()),
        "proporciones": list(mortero.PROPORCIONES.keys()),
        "diams_estribo": [{"nombre": k, "mm": v} for k, v in estribos.DIAMS_ESTRIBO.items()],
        "diams_long": [{"nombre": k, "mm": v} for k, v in estribos.DIAMS_LONG.items()],
        "tipos_elemento": [
            {
                "clave": t.clave,
                "nombre": t.nombre,
                "ayuda": t.ayuda,
                "campos": [
                    {"clave": c.clave, "etiqueta": c.etiqueta, "valor": c.valor,
                     "sufijo": c.sufijo, "paso": c.paso, "decimales": c.decimales,
                     "minimo": c.minimo, "maximo": c.maximo}
                    for c in t.campos
                ],
            }
            for t in elementos.TIPOS
        ],
        "tabla_baldes": [
            {
                "fc": row[0], "proporcion": row[4],
                "b18_arena": round(baldes.parse_proporcion(row[4])[1] * 0.0283168 / 0.018, 2),
                "b18_piedra": round(baldes.parse_proporcion(row[4])[2] * 0.0283168 / 0.018, 2),
                "b20_arena": round(baldes.parse_proporcion(row[4])[1] * 0.0283168 / 0.020, 2),
                "b20_piedra": round(baldes.parse_proporcion(row[4])[2] * 0.0283168 / 0.020, 2),
            }
            for row in modelo.TABLA_BASE
        ],
        "peso_varillas": [
            {"nombre": k, "mm": v["mm"], "kg_m": v["kg_m"], "largo_m": v["largo_m"],
             "kg_varilla": round(v["kg_m"] * v["largo_m"], 2)}
            for k, v in elementos.PESO_VARILLA.items()
        ],
    })


if __name__ == "__main__":
    print("\n  Metrados · Constructor IA")
    print("  http://localhost:5050\n")
    app.run(host="0.0.0.0", port=5050, debug=False)
