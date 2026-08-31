"""Exporta un requerimiento de dosificacion a un Excel con formato, listo para expediente."""

from pathlib import Path
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

AZUL = "6C3CE0"
CIAN = "12C9C2"
GRIS = "F2F2F2"

ASSETS = Path(__file__).resolve().parent / "assets"


def _insertar_logo(ws):
    ruta = ASSETS / "constructor_ia_logo.png"
    if not ruta.exists():
        return
    try:
        from openpyxl.drawing.image import Image as XLImage

        img = XLImage(str(ruta))
        ratio = img.width / img.height if img.height else 3.0
        img.height = 44
        img.width = int(44 * ratio)
        ws.add_image(img, "C1")
    except Exception:
        return


def exportar(ruta, req, responsable="", presupuesto=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dosificacion"
    _insertar_logo(ws)

    borde = Border(*(Side(style="thin", color="BFBFBF"),) * 4)
    titulo_font = Font(bold=True, size=14, color="FFFFFF")
    subtitulo_font = Font(bold=True, size=10, color="595959")
    encabezado_fill = PatternFill("solid", fgColor=AZUL)
    encabezado_font = Font(bold=True, color="FFFFFF")
    seccion_fill = PatternFill("solid", fgColor=GRIS)

    ws.merge_cells("A1:D1")
    ws["A1"] = "MEMORIA DE DOSIFICACION DE CONCRETO"
    ws["A1"].font = titulo_font
    ws["A1"].fill = PatternFill("solid", fgColor=AZUL)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws["A2"] = "Constructor IA · Metodo: dosificacion por volumenes (tabla referencial ACI / practica peruana)"
    ws["A2"].font = subtitulo_font
    ws.merge_cells("A2:D2")

    fila = 4
    ws.cell(row=fila, column=1, value="Datos de entrada").font = encabezado_font
    ws.cell(row=fila, column=1).fill = encabezado_fill
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
    fila += 1

    datos_entrada = [
        ("f'c (kg/cm2)", req.fc),
        ("Volumen a producir (m3)", req.volumen_m3),
    ]
    for etiqueta, valor in datos_entrada:
        ws.cell(row=fila, column=1, value=etiqueta)
        ws.cell(row=fila, column=2, value=valor)
        fila += 1

    fila += 1
    ws.cell(row=fila, column=1, value="Parametros de dosificacion (por m3)").font = encabezado_font
    ws.cell(row=fila, column=1).fill = encabezado_fill
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
    fila += 1

    dos = req.dosificacion
    parametros = [
        ("Relacion a/c", dos.a_c),
        ("Slump (pulgadas)", dos.slump_pulg),
        ("Tamano max. agregado (pulg)", dos.tmax_pulg),
        ("Dosificacion en volumen", dos.dosificacion_volumen),
        ("Interpolado", "Si" if dos.interpolado else "No"),
    ]
    for etiqueta, valor in parametros:
        ws.cell(row=fila, column=1, value=etiqueta)
        ws.cell(row=fila, column=2, value=valor)
        fila += 1

    fila += 1
    ws.cell(row=fila, column=1, value="Material").font = encabezado_font
    ws.cell(row=fila, column=2, value="Cantidad total").font = encabezado_font
    ws.cell(row=fila, column=3, value="Unidad").font = encabezado_font
    for c in (1, 2, 3):
        ws.cell(row=fila, column=c).fill = encabezado_fill
    fila_tabla_inicio = fila + 1

    materiales = [
        ("Cemento", req.cemento_bolsas, "bolsas (42.5 kg)"),
        ("Cemento", req.cemento_kg, "kg"),
        ("Arena", req.arena_m3, "m3"),
        ("Piedra / agregado grueso", req.piedra_m3, "m3"),
        ("Agua", req.agua_m3, "m3"),
        ("Agua", req.agua_litros, "litros"),
    ]
    fila += 1
    for nombre, valor, unidad in materiales:
        ws.cell(row=fila, column=1, value=nombre)
        ws.cell(row=fila, column=2, value=valor)
        ws.cell(row=fila, column=3, value=unidad)
        fila += 1

    for r in range(fila_tabla_inicio - 1, fila):
        for c in (1, 2, 3):
            ws.cell(row=r, column=c).border = borde

    if presupuesto is not None:
        fila += 1
        titulos = ["Material", "Cantidad", "P.U. (S/)", "Parcial (S/)"]
        for c, t in enumerate(titulos, start=1):
            cel = ws.cell(row=fila, column=c, value=t)
            cel.font = encabezado_font
            cel.fill = encabezado_fill
        pres_inicio = fila
        fila += 1
        for ln in presupuesto.lineas:
            ws.cell(row=fila, column=1, value=ln.material)
            ws.cell(row=fila, column=2, value=f"{ln.cantidad:,.2f} {ln.unidad}")
            ws.cell(row=fila, column=3, value=round(ln.precio_unit, 2))
            ws.cell(row=fila, column=4, value=round(ln.parcial, 2))
            fila += 1

        cel_tot = ws.cell(row=fila, column=1, value="TOTAL materiales")
        cel_tot.font = Font(bold=True)
        cel_val = ws.cell(row=fila, column=4, value=round(presupuesto.total, 2))
        cel_val.font = Font(bold=True, color="FFFFFF")
        for c in (1, 2, 3, 4):
            ws.cell(row=fila, column=c).fill = PatternFill("solid", fgColor=AZUL)
            ws.cell(row=fila, column=c).font = Font(bold=True, color="FFFFFF")
        fila += 1
        ws.cell(row=fila, column=1, value="Costo por m3 de concreto")
        ws.cell(row=fila, column=4, value=round(presupuesto.costo_m3, 2))
        for r in range(pres_inicio, fila + 1):
            for c in (1, 2, 3, 4):
                ws.cell(row=r, column=c).border = borde

    fila += 2
    ws.cell(row=fila, column=1, value=f"Fecha: {date.today().isoformat()}")
    fila += 1
    ws.cell(row=fila, column=1, value=f"Responsable: {responsable or '-'}")

    ws.column_dimensions[get_column_letter(1)].width = 32
    ws.column_dimensions[get_column_letter(2)].width = 22
    ws.column_dimensions[get_column_letter(3)].width = 18
    ws.column_dimensions[get_column_letter(4)].width = 14

    wb.save(ruta)


def exportar_metrado(ruta, elementos, total_m3, responsable=""):
    """Exporta la lista de elementos metrados a Excel con formato para expediente."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Metrado"
    _insertar_logo(ws)

    borde = Border(*(Side(style="thin", color="BFBFBF"),) * 4)
    titulo_font = Font(bold=True, size=14, color="FFFFFF")
    encabezado_font = Font(bold=True, color="FFFFFF")
    encabezado_fill = PatternFill("solid", fgColor=AZUL)
    total_fill = PatternFill("solid", fgColor="12C9C2")

    ws.merge_cells("A1:E1")
    ws["A1"] = "METRADO DE CONCRETO - ELEMENTOS ESTRUCTURALES"
    ws["A1"].font = titulo_font
    ws["A1"].fill = PatternFill("solid", fgColor=AZUL)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws["A2"] = "Constructor IA · Volumen de concreto por elemento"
    ws.merge_cells("A2:E2")

    fila = 4
    encabezados = ["Item", "Elemento", "Dimensiones", "Cant.", "Vol. (m3)"]
    for c, h in enumerate(encabezados, start=1):
        cel = ws.cell(row=fila, column=c, value=h)
        cel.font = encabezado_font
        cel.fill = encabezado_fill
        cel.border = borde
        cel.alignment = Alignment(horizontal="center")
    fila += 1

    inicio_tabla = fila
    for i, e in enumerate(elementos, start=1):
        ws.cell(row=fila, column=1, value=i).alignment = Alignment(horizontal="center")
        ws.cell(row=fila, column=2, value=e.nombre)
        ws.cell(row=fila, column=3, value=e.descripcion)
        ws.cell(row=fila, column=4, value=e.cantidad).alignment = Alignment(horizontal="center")
        vol = ws.cell(row=fila, column=5, value=round(e.volumen_total, 4))
        vol.alignment = Alignment(horizontal="right")
        for c in range(1, 6):
            ws.cell(row=fila, column=c).border = borde
        fila += 1

    # fila de total
    ws.cell(row=fila, column=4, value="TOTAL").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=fila, column=4).fill = total_fill
    ws.cell(row=fila, column=5, value=round(total_m3, 3)).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=fila, column=5).fill = total_fill
    ws.cell(row=fila, column=5).alignment = Alignment(horizontal="right")
    for c in range(1, 6):
        ws.cell(row=fila, column=c).border = borde
    fila += 2

    ws.cell(row=fila, column=1, value=f"Fecha: {date.today().isoformat()}")
    fila += 1
    ws.cell(row=fila, column=1, value=f"Responsable: {responsable or '-'}")

    ws.column_dimensions[get_column_letter(1)].width = 6
    ws.column_dimensions[get_column_letter(2)].width = 28
    ws.column_dimensions[get_column_letter(3)].width = 38
    ws.column_dimensions[get_column_letter(4)].width = 8
    ws.column_dimensions[get_column_letter(5)].width = 14

    wb.save(ruta)
